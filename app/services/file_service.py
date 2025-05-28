import shutil
import uuid
from pathlib import Path
from typing import Tuple, List, Optional, Union, Dict, Any
from datetime import datetime, timezone
import mimetypes # For better content type detection
import os
import traceback

from fastapi import UploadFile, HTTPException, status
import pandas as pd
import openpyxl # Or another library like openpyxl directly if preferred

from app.core.config import settings

# Define allowed Excel MIME types and extensions
ALLOWED_EXCEL_MIME_TYPES = [
    "application/vnd.ms-excel",  # .xls
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.oasis.opendocument.spreadsheet", # .ods (optional, if you want to support)
]
ALLOWED_EXTENSIONS = [".xls", ".xlsx", ".ods"] # .ods is optional

# In-memory store for uploaded file metadata
# Ensure Dict and Any are imported from typing for this
FILE_METADATA: Dict[str, Dict[str, Any]] = {} # Added this line to define FILE_METADATA

# 配置上传目录
UPLOAD_DIR = Path("app/temp_files")

async def save_uploaded_file(file: UploadFile, file_id: str) -> Path:
    """
    保存上传的文件到临时目录
    """
    try:
        # 确保上传目录存在
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        
        # 获取文件扩展名
        file_extension = Path(file.filename).suffix if file.filename else ".xlsx"
        
        # 构建目标文件路径
        destination_filename = f"{file_id}{file_extension}"
        destination_path = UPLOAD_DIR / destination_filename
        
        # 保存文件
        with destination_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # 获取文件大小（KB）
        file_size_kb = destination_path.stat().st_size / 1024
        print(f"[file_service] Saved file: id={file_id}, original_name='{file.filename}', stored_path='{destination_path}', size_kb={file_size_kb:.2f}")
        
        return destination_path

    except Exception as e:
        print(f"[file_service] Error saving file: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to save uploaded file: {str(e)}"
        )

async def get_file_path(file_id: str) -> Path:
    """
    获取已保存文件的路径
    """
    try:
        # 尝试查找文件（支持多种可能的扩展名）
        possible_extensions = [".xlsx", ".xls"]
        for ext in possible_extensions:
            file_path = UPLOAD_DIR / f"{file_id}{ext}"
            if file_path.exists() and file_path.is_file():
                print(f"[file_service] For file_id '{file_id}': retrieved stored_path_str='{file_path}', Path_object='{file_path}', Path_is_absolute={file_path.is_absolute()}")
                print(f"[file_service] For file_id '{file_id}', path '{file_path}': exists() -> {file_path.exists()}, is_file() -> {file_path.is_file()}")
                return file_path
                
        raise FileNotFoundError(f"No file found for ID '{file_id}' with supported extensions.")
        
    except FileNotFoundError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )
    except Exception as e:
        print(f"[file_service] Error getting file path: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error accessing file with ID '{file_id}': {str(e)}"
        )

# --- Excel specific functions (to be expanded) ---

async def read_excel_column(
    file_path: Path,
    column_identifier: str,
    sheet_name: Optional[Union[str, int]] = 0 # Allow int for sheet index as well, default to first sheet
) -> List[str]:
    """
    Reads a specific column from an Excel file into a list of strings.

    Args:
        file_path: Path to the Excel file.
        column_identifier: The column to read. Can be a column letter (e.g., 'A'),
                           a header name (e.g., 'Source Text'), or a 0-indexed integer.
        sheet_name: Name or index of the sheet to read from. Defaults to the first sheet (index 0).

    Returns:
        A list of strings from the specified column.

    Raises:
        HTTPException: If the file cannot be processed or the column is not found.
    """
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Excel file not found at path: {file_path}")

    engine = None
    if file_path.suffix.lower() in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        engine = 'openpyxl'
    elif file_path.suffix.lower() == '.xls':
        engine = 'xlrd' # or 'openpyxl' if xlrd is not preferred/installed for .xls
    # For .ods, pandas might use 'odf'

    try:
        # Pandas can directly handle column letters in usecols.
        # It can also handle integer indices for columns if header=None or if the index is of a valid column name.
        # If column_identifier is intended as a name, it must match a header.
        
        # First, try to see if it's a valid header name by reading only headers
        # This helps in providing a better error message if it's neither a name nor a valid letter/index
        try:
            df_check = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0, engine=engine)
            is_header_name = column_identifier in df_check.columns
        except Exception:
            # If reading headers fails (e.g. empty sheet, password protected without handling, etc.)
            # proceed with caution, pandas will try its best with the identifier.
            is_header_name = False
            df_check = None # No columns to list in case of early failure

        # Let pandas try to resolve column_identifier (name, letter, or int index if applicable)
        # If it's an integer string like "0", "1", convert to int for usecols if it's not a header name.
        usecols_arg = column_identifier
        if not is_header_name and column_identifier.isdigit():
            try:
                # Try to convert to int, assuming it's a 0-based index if not a header name.
                # However, pandas' usecols usually expects string names or direct int indices.
                # If it's a string like '0', pandas might interpret it as a name '0'.
                # For safety, if it is purely digit and not a header, we might treat it as index.
                # But pandas is quite flexible with string column letters ('A', 'B') directly in usecols.
                # So, direct usage of column_identifier is often fine.
                pass # Let pandas try to interpret it as is. It's good with 'A', 'B'. If it's '0', it may be a name.
            except ValueError:
                pass # Not a simple integer string

        df = pd.read_excel(
            file_path, 
            sheet_name=sheet_name, 
            usecols=[usecols_arg] if is_header_name else usecols_arg, # Pass as list if known header, else as is for pandas to try
            header=0, # Assume header is at the first row for name resolution
            keep_default_na=False, 
            dtype=str,
            engine=engine
        )
        
        if df.empty:
            # If df is empty, it could be that the column exists but has no data, 
            # or that usecols resolved to an empty set (e.g. valid column letter but no data there)
            # If the column *name* or *letter* was invalid, pandas would likely have raised an error already.
            # We can add a check here: if df_check exists, verify column_identifier was valid.
            if df_check is not None and not is_header_name:
                 # If it wasn't a header and it's not a digit (potential index), 
                 # and it's not a typical column letter pattern, it might be an issue.
                 # However, pandas' own error for invalid usecols is usually sufficient.
                 # This path (df.empty) after a successful read_excel call (no error) typically means
                 # the specified column(s) were found but contained no data rows, or only NaN-like values
                 # which keep_default_na=False + dtype=str should handle by returning empty strings.
                 pass # It's fine, column was found but is empty or has no actual data rows
            return [] 
        
        # If usecols resulted in multiple columns (e.g. if a non-list was passed and pandas interpreted it broadly),
        # ensure we only take the first one that matches or is implied.
        # However, with usecols=[col_name] or usecols='A', it should return only one.
        if column_identifier in df.columns:
            actual_column_to_extract = column_identifier
        elif df.columns:
            actual_column_to_extract = df.columns[0]
        else: # Should not happen if df is not empty
             raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Column '{column_identifier}' could not be resolved in sheet '{sheet_name}'.")

        column_data = df[actual_column_to_extract].astype(str).tolist()
        return column_data

    except FileNotFoundError: # This is caught by the check at the beginning now
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Excel file not found at path: {file_path}")
    except ValueError as ve:
        detail_msg = f"Invalid column identifier '{column_identifier}' or sheet '{sheet_name}'."
        if df_check is not None and df_check.columns is not None:
            detail_msg += f" Available columns: {df_check.columns.tolist()}. Error: {ve}"
        else:
            detail_msg += f" Error: {ve}"
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=detail_msg)
    except KeyError as ke: # Raised by pandas if a column name in usecols list is not found
        detail_msg = f"Column name '{column_identifier}' not found in sheet '{sheet_name}'."
        if df_check is not None and df_check.columns is not None:
            detail_msg += f" Available columns: {df_check.columns.tolist()}. Error: {ke}"
        else:
            detail_msg += f" Error: {ke}"
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=detail_msg)
    except Exception as e:
        # Log error e for server-side debugging
        print(f"Pandas Read Error for column '{column_identifier}' in sheet '{sheet_name}': {type(e).__name__} - {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Could not read column '{column_identifier}' from Excel file. Ensure it's a valid Excel file and the column/sheet exists.")

# Example of how to get file_id without extension, if needed for some internal logic
# def get_file_id_from_filename(filename_with_extension: str) -> str:
#     return Path(filename_with_extension).stem 

async def write_excel_column(
    original_file_path: Path,
    texts_to_write: List[str],
    new_column_name: str,
    output_file_id: str, # This will be used to name the output file, e.g., job_id
    sheet_name: Optional[str] = None # Optional: specify sheet name, defaults to first sheet
) -> Path:
    """
    Reads an existing Excel file, adds a new column with the provided texts,
    and saves it as a new file in the TEMP_FILES_DIR.

    Args:
        original_file_path: Path to the original Excel file.
        texts_to_write: List of strings to write into the new column.
        new_column_name: The header name for the new column.
        output_file_id: The ID to use for naming the output file (e.g., job_id).
                        The output file will be named `output_file_id.xlsx`.
        sheet_name: Optional name of the sheet to operate on. Defaults to the first sheet.

    Returns:
        Path to the newly created Excel file with the translated text.

    Raises:
        HTTPException if there are issues reading the original file or writing the new file.
    """
    if not original_file_path.exists() or not original_file_path.is_file():
        print(f"Error: Original file not found at: {original_file_path}")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Original file for processing not found at path: {original_file_path}"
        )

    output_filename = f"{output_file_id}.xlsx"
    # Ensure settings.TEMP_FILES_DIR is Path object and correctly configured
    output_file_path = settings.TEMP_FILES_DIR / output_filename

    try:
        # Read the original Excel file
        if sheet_name:
            df = pd.read_excel(original_file_path, sheet_name=sheet_name, engine='openpyxl')
        else:
            df = pd.read_excel(original_file_path, engine='openpyxl') # Reads the first sheet by default

        # Add the new column with the texts to write
        # Ensure the list of texts matches the number of rows in the DataFrame
        # If not, pad with empty strings or raise an error, depending on desired behavior.
        # For now, we'll assume it can be shorter and pandas will handle it (filling NaNs for shorter lists).
        # If it's longer, it might cause issues or be truncated.
        # A more robust solution would be to align lengths or handle mismatches explicitly.
        if len(texts_to_write) > len(df):
            print(f"Warning: Number of texts to write ({len(texts_to_write)}) exceeds number of rows in Excel ({len(df)}). Truncating.")
            df[new_column_name] = texts_to_write[:len(df)]
        elif len(texts_to_write) < len(df):
            print(f"Warning: Number of texts to write ({len(texts_to_write)}) is less than number of rows in Excel ({len(df)}). Padding with empty strings.")
            padded_texts = texts_to_write + [''] * (len(df) - len(texts_to_write))
            df[new_column_name] = padded_texts
        else:
            df[new_column_name] = texts_to_write

        # Save the modified DataFrame to a new Excel file
        # Make sure the TEMP_FILES_DIR exists (should be handled by settings initialization)
        os.makedirs(settings.TEMP_FILES_DIR, exist_ok=True)
        df.to_excel(output_file_path, index=False, engine='openpyxl')

        print(f"Info: Successfully wrote translated texts to new column '{new_column_name}' in file: {output_file_path}")
        return output_file_path

    except FileNotFoundError:
        print(f"Error: Pandas could not find the original file during read: {original_file_path}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to read the original Excel file at: {original_file_path}"
        )
    except ValueError as ve:
        print(f"Error: ValueError during Excel processing (e.g., sheet not found '{sheet_name}'): {ve} for file {original_file_path}")
        detail_msg = f"Error processing Excel sheet '{sheet_name}' in file {original_file_path.name}. Ensure the sheet exists."
        if not sheet_name:
            detail_msg = f"Error processing Excel file {original_file_path.name}. It might be empty or corrupted."
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=detail_msg)
    except Exception as e:
        print(f"Error: An unexpected error occurred while writing to Excel column for output ID '{output_file_id}': {e}")
        traceback.print_exc() # Log full traceback for server-side debugging
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"An unexpected error occurred while generating the translated Excel file."
        )

# Placeholder for cleaning up old files if necessary (e.g., a background task)
# async def cleanup_temp_file(file_path: Path):

async def write_excel_with_translations(
    original_file_path: Path,
    translations: List[str],
    original_text_column_name: str, # Name or letter of the original text column
    new_translated_column_name: str, # Desired name for the new column with translations
    project_name: Optional[str] = None,
    base_filename_suffix: str = "_translated"
) -> Path:
    """
    Reads an original Excel file, adds a new column with translated texts,
    and saves it as a new Excel file.

    Args:
        original_file_path: Path to the source Excel file.
        translations: A list of translated strings, in the same order as original texts.
        original_text_column_name: Identifier for the column containing original text.
        new_translated_column_name: Name for the new column for translations.
        project_name: Optional project name to include in the output filename.
        base_filename_suffix: Suffix to add to the original filename for the output file.

    Returns:
        Path to the newly created Excel file with translations.
    
    Raises:
        FileNotFoundError: If the original file does not exist.
        ValueError: If the original text column is not found or other input issues.
        Exception: For other pandas/Excel processing errors.
    """
    print(f"[file_service] Attempting to write translations to Excel. Original: {original_file_path}")
    if not original_file_path.exists():
        raise FileNotFoundError(f"Original file not found at: {original_file_path}")

    try:
        # Read the Excel file into a pandas DataFrame
        # We need to handle both column names and column letters (e.g., 'A', 'B')
        # For simplicity, let's assume original_text_column_name is a header name for now.
        # If it can be a letter, more complex logic is needed to find the header or use index.
        
        # df = pd.read_excel(original_file_path, sheet_name=0) # Read the first sheet
        # A more robust way is to load the workbook and work with the active sheet
        # or allow specifying a sheet name/index.
        
        # Using openpyxl directly for more control, especially if dealing with existing formatting
        workbook = openpyxl.load_workbook(original_file_path)
        sheet = workbook.active # Get the active sheet

        # Find the original text column
        original_col_idx = None
        # Try to find by header name first
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == original_text_column_name:
                original_col_idx = col
                break
        
        # If not found by header, and if original_text_column_name looks like a letter (e.g. "A", "B")
        if original_col_idx is None and isinstance(original_text_column_name, str) and original_text_column_name.isalpha() and len(original_text_column_name) <= 2 : # Basic check for column letter
             try:
                 original_col_idx = openpyxl.utils.column_index_from_string(original_text_column_name.upper())
             except Exception: # openpyxl.utils.exceptions.IllegalCharacterError
                 pass # Could not convert to index

        if original_col_idx is None:
            raise ValueError(f"Original text column '{original_text_column_name}' not found in the Excel sheet '{sheet.title}'.")

        # Find the first empty column for the new translated text, or append after the last column
        # Or, place it right next to the original column if desired.
        # For simplicity, let's add it as a new last column or one after original_col_idx.
        # Let's try to place it one column after the original text column
        # Ensure this new column index doesn't overwrite existing data unintentionally if it's not the very last.
        
        # A safer approach for new column: find first completely empty column or append
        # For now, let's just put it after the current max_column
        translated_col_idx = sheet.max_column + 1
        sheet.cell(row=1, column=translated_col_idx, value=new_translated_column_name) # Set header for new column

        # Write the translations
        # Assuming translations list matches row-by-row with original texts, starting from row 2 (after header)
        if len(translations) > sheet.max_row -1 : # If more translations than rows (excluding header)
             print(f"[file_service] Warning: Number of translations ({len(translations)}) is greater than data rows in Excel ({sheet.max_row -1}). Some translations might be ignored.")
        
        for i, translated_text in enumerate(translations):
            # Excel rows are 1-indexed, data starts from row 2 if header is in row 1
            target_row = i + 2 
            if target_row > sheet.max_row: # Don't write past existing data rows if translations are too many
                # Or, append new rows if that's desired behavior.
                # For now, only write up to existing max_row (matching original data length)
                # This assumes original_texts and translations have same length
                # and this length matches the number of data rows in original_col_idx
                # A more robust approach would align based on the number of original texts read.
                break 
            sheet.cell(row=target_row, column=translated_col_idx, value=translated_text)

        # Construct the output file path
        output_dir = settings.OUTPUT_FILES_DIR # Make sure this is configured
        if not output_dir.exists():
            output_dir.mkdir(parents=True, exist_ok=True)

        base_name = original_file_path.stem
        extension = original_file_path.suffix
        
        filename_parts = [base_name]
        if project_name:
            filename_parts.append(project_name.replace(" ", "_")) # Sanitize project name for filename
        filename_parts.append(base_filename_suffix)
        
        output_filename = "_".join(filter(None, filename_parts)) + extension
        output_file_path = output_dir / output_filename
        
        # Ensure filename is unique if it already exists to avoid overwriting (optional)
        # counter = 1
        # temp_path = output_file_path
        # while temp_path.exists():
        #     output_filename = "_".join(filter(None, filename_parts)) + f"_({counter})" + extension
        #     temp_path = output_dir / output_filename
        #     counter += 1
        # output_file_path = temp_path

        workbook.save(output_file_path)
        print(f"[file_service] Translated Excel file saved to: {output_file_path}")
        return output_file_path

    except FileNotFoundError:
        raise
    except ValueError: # For column not found
        raise
    except Exception as e:
        print(f"[file_service] Error processing Excel file {original_file_path}: {type(e).__name__} - {e}")
        traceback.print_exc()
        raise Exception(f"Failed to write translated Excel file: {e}") # Re-raise as a general exception